# plf.py
# ----------
#
import boto3
import datetime
import json
import openpyxl
import pystache
import re
import shutil
import sys
import time
import yaml

if len(sys.argv) != 3:
    raise Exception('Usage: python3 plf.py [AMI_ID] [TEMPLATE_VERSION]')

OE_MARKUP_PERCENTAGE = 0.05
ANNUAL_SAVINGS_PERCENTAGE = 0.80 # 20% off
MINIMUM_RATE = 0.01
HOURS_IN_A_YEAR = 8760
DEFAULT_REGION = 'us-east-1'
AMI=sys.argv[1]
VERSION=sys.argv[2]

pricing = boto3.client('pricing', region_name=DEFAULT_REGION)

regions = {
    'AWS GovCloud (US-East)'    :'us-gov-east-1',
    'AWS GovCloud (US-West)'    :'us-gov-west-1',
    'Africa (Cape Town)'        :'af-south-1',
    'Asia Pacific (Hong Kong)'  :'ap-east-1',
    'Asia Pacific (Mumbai)'     :'ap-south-1',
    'Asia Pacific (Osaka)'      :'ap-northeast-3',
    'Asia Pacific (Seoul)'      :'ap-northeast-2',
    'Asia Pacific (Singapore)'  :'ap-southeast-1',
    'Asia Pacific (Sydney)'     :'ap-southeast-2',
    'Asia Pacific (Tokyo)'      :'ap-northeast-1',
    'Canada (Central)'          :'ca-central-1',
    'China (Beijing)'           :'none',
    'China (Ningxia)'           :'none',
    'EU (Frankfurt)'            :'eu-central-1',
    'EU (Ireland)'              :'eu-west-1',
    'EU (London)'               :'eu-west-2',
    'EU (Milan)'                :'eu-south-1',
    'EU (Paris)'                :'eu-west-3',
    'EU (Stockholm)'            :'eu-north-1',
    'Middle East (Bahrain)'     :'me-south-1',
    'South America (Sao Paulo)' :'sa-east-1',
    'US East (Miami)'           :'none',
    'US East (N. Virginia)'     :'us-east-1',
    'US East (Ohio)'            :'us-east-2',
    'US West (Los Angeles)'     :'none',
    'US West (N. California)'   :'us-west-1',
    'US West (Oregon)'          :'us-west-2'
}

plf_config = yaml.load(
    open('/code/plf_config.yaml'),
    Loader=yaml.SafeLoader
)
allowed_values = yaml.load(
    open('/code/allowed_values.yaml'),
    Loader=yaml.SafeLoader
)
allowed_instance_types = allowed_values['allowed_instance_types']
allowed_regions = open('/code/supported_regions.txt').read().split('\n')

def get_highest_hourly_price_for_instance_type(instance_type, allowed_regions):
    highest_hourly_price = 0
    highest_hourly_region = None
    response = pricing.get_products(
        ServiceCode='AmazonEC2',
        Filters = [
            {'Type' :'TERM_MATCH', 'Field':'capacitystatus',  'Value':'Used' },
            {'Type' :'TERM_MATCH', 'Field':'instanceType',    'Value':instance_type },
            {'Type' :'TERM_MATCH', 'Field':'licenseModel',    'Value':'No License required' },
            {'Type' :'TERM_MATCH', 'Field':'operatingSystem', 'Value':'Linux' },
            {'Type' :'TERM_MATCH', 'Field':'preInstalledSw',  'Value':'NA' },
            {'Type' :'TERM_MATCH', 'Field':'tenancy',         'Value':'Shared' },
            {'Type' :'TERM_MATCH', 'Field':'termType',        'Value':'OnDemand' }
        ],
        MaxResults=100
    )
    for price in response['PriceList']:
        priceObj = json.loads(price)
        location = priceObj['product']['attributes']['location']
        if location in regions and regions[location] in allowed_regions:
            termsKey = next(iter(priceObj['terms']['OnDemand']))
            priceDimensionsKey = next(iter(priceObj['terms']['OnDemand'][termsKey]['priceDimensions']))
            hourly_price = float(priceObj['terms']['OnDemand'][termsKey]['priceDimensions'][priceDimensionsKey]['pricePerUnit']['USD'])
            # print(f'Hourly price for {location} is {hourly_price}')
            if hourly_price > highest_hourly_price:
                highest_hourly_price = hourly_price
                highest_hourly_region = location
    # print(f'Highest price for {instance_type} is ${highest_hourly_price} at {highest_hourly_region}')
    return highest_hourly_price

src = 'plf.xlsx'
now_dt = datetime.datetime.now()
dst = f"plf-version-{VERSION.replace('.', '-')}--gen-{now_dt.strftime('%Y%m%d-%H%M%S')}.xlsx"
SHEET_NAME = 'SSLSingleAMIAndCARWithContract'

shutil.copyfile(src, dst)

dst_wb = openpyxl.load_workbook(dst)
dst_sheet = dst_wb[SHEET_NAME]

src_wb = openpyxl.load_workbook(src)
src_sheet = src_wb[SHEET_NAME]
headers = src_sheet[5]
values = src_sheet[6]

current_column_index = 0
for header in headers:
    column = header.value
    value = ''
    availability_match = re.search(r'(.+) Availability', column)
    if availability_match:
        match_keyword = availability_match.groups()[0]
        # region or instance availability?
        is_instance_match = re.search(r'^(.+)\.(.+)$', match_keyword)
        if is_instance_match:
            if match_keyword in allowed_instance_types:
                value = 'TRUE'
            else:
                value = ''
        else:
            if match_keyword in allowed_regions:
                value = 'TRUE'
            else:
                value = ''

    price_match = re.search(r'(.+) (Hourly|Annual) Price', column)
    if price_match:
        instance_type = price_match.groups()[0]
        if instance_type in allowed_instance_types:
            price_type = price_match.groups()[1]
            price = get_highest_hourly_price_for_instance_type(instance_type, allowed_regions)
            hourly_price_with_markup = round(price * OE_MARKUP_PERCENTAGE, 2)
            if hourly_price_with_markup < MINIMUM_RATE:
                hourly_price_with_markup = MINIMUM_RATE
            if price_type == 'Hourly':
                value = '{:.3f}'.format(hourly_price_with_markup)
            else:
                annual_price = hourly_price_with_markup * HOURS_IN_A_YEAR * ANNUAL_SAVINGS_PERCENTAGE
                value = '{:.3f}'.format(round(annual_price, 2))
    if not availability_match and not price_match:
        if column in plf_config:
            value = pystache.render(plf_config[column], {'ami': AMI, 'version': VERSION})
        else:
            value = None
    if value is not None and value != values[current_column_index].value:
        print(f"{column} has changed! Old: '{values[current_column_index].value}' New: '{value}'")
        dst_sheet.cell(row=6, column=current_column_index+1, value=value)
    current_column_index += 1

dst_wb.save(dst)
print(f'PLF saved to {dst}')
